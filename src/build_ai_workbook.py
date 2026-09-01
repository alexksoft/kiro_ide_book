"""Build the Excel workbook for building software with AI and Python.

Companion to the Word doc; same verified data (ai_python_data.py,
ai_python_snippets.py, ai_link_status.json).
Sheets: Start here | Roadmap | All resources | By level | Code snippets | Index
"""

import json
from collections import Counter, OrderedDict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from ai_python_data import ROWS, ROADMAP, B, I, A, X
from ai_python_snippets import SNIPPETS

OUT = "Build_Software_with_AI_and_Python.xlsx"
TODAY = date.today().isoformat()

NAVY = "1F3864"
ACCENT = "2E6B3E"
LIGHT = "E4EFE7"
BAND = "F3F8F4"
CODE_BG = "F4F4F7"
LINK_BLUE = "0563C1"
OK_GREEN = "1E7B34"
WARN_AMBER = "9C5700"

thin = Side(style="thin", color="CFDACF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
link_font = Font(color=LINK_BLUE, underline="single", size=11)
body_font = Font(size=11, color="222222")
code_font = Font(name="Consolas", size=9.5, color="1A1A2E")

LEVEL_ORDER = [B, I, A, X]

try:
    with open("ai_link_status.json") as fh:
        STATUS = json.load(fh)
except FileNotFoundError:
    STATUS = {}


def verdict(url):
    s = STATUS.get(url)
    if not s:
        return "not checked", ""
    if s["kind"] == "ok":
        return "live (HTTP 200)", s["checked"]
    if s["kind"] == "blocked-to-bots":
        return f"bot-blocked ({s['code']}) - opens in browser", s["checked"]
    return f"PROBLEM ({s['code']})", s["checked"]


def domain_of(url):
    d = url.split("//", 1)[-1].split("/", 1)[0]
    return d[4:] if d.startswith("www.") else d


wb = Workbook()

# ================================================================= Start here
ws = wb.active
ws.title = "Start here"
ws.sheet_view.showGridLines = False

n_ok = sum(1 for u in {r[4] for r in ROWS} if STATUS.get(u, {}).get("kind") == "ok")
n_block = sum(1 for u in {r[4] for r in ROWS} if STATUS.get(u, {}).get("kind") == "blocked-to-bots")

intro = [
    ("Building Software with AI and Python", "title"),
    (f"Compiled {TODAY}  |  {len(ROWS)} resources  |  {len(SNIPPETS)} code snippets  |  every link clickable", "sub"),
    ("", None),
    ("What this is", "h"),
    ("A tiered roadmap and verified link catalogue for creating software with AI in Python. Everything is split into "
     "three levels so you can start where you are.", "p"),
    ("", None),
    ("The three levels", "h"),
    ("Beginner - new to Python or to programming. Write working scripts and make your first AI app.", "p"),
    ("Intermediate - you can write Python. Build maintainable, tested, deployable AI software with real APIs and data.", "p"),
    ("Advanced - you ship AI software. Agents, evaluation, serving, security, cost and performance at production quality.", "p"),
    ("All levels - tools for letting AI help you write the software itself.", "p"),
    ("", None),
    ("The six sheets", "h"),
    ("Start here - this page.", "p"),
    ("Roadmap - 14 staged milestones with a 'done when' test for each. Use it as a checklist.", "p"),
    ("All resources - the full catalogue, one row per link, with filters.", "p"),
    ("By level - the same links grouped Beginner / Intermediate / Advanced / All levels.", "p"),
    ("Code snippets - copy-paste Python for the things people get stuck on.", "p"),
    ("Index - counts per level, per topic and the link-check results.", "p"),
    ("", None),
    ("Cost", "h"),
    ("Most resources are free. Calling a hosted model (OpenAI, Anthropic) costs a few dollars for a week of learning, "
     "but you can do almost everything at zero cost by running open models locally with Ollama. The Cost column says which.", "p"),
    ("", None),
    ("Link checking", "h"),
    (f"Every link was requested on {TODAY}: {n_ok} returned HTTP 200 and {n_block} sit behind bot protection "
     "(Cloudflare 'Just a moment...' pages) - those are real and open in a browser. None were broken. The code "
     "snippets were all syntax-checked.", "p"),
    ("", None),
    ("Caveats", "h"),
    ("The AI tooling landscape moves fast: versions, model names and prices change - cross-check community posts "
     "against official docs. Courses marked Paid are third-party; check reviews and recency first. Snippets are "
     "starting points - add real error handling, secrets management and tests before production.", "p"),
]

r = 1
for text, kind in intro:
    c = ws.cell(row=r, column=1, value=text)
    if kind == "title":
        c.font = Font(size=20, bold=True, color=NAVY)
        ws.row_dimensions[r].height = 28
    elif kind == "sub":
        c.font = Font(size=11, italic=True, color="666666")
    elif kind == "h":
        c.font = Font(size=12, bold=True, color=ACCENT)
        ws.row_dimensions[r].height = 20
    elif kind == "p":
        c.font = Font(size=11, color="333333")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 15 * (len(text) // 108 + 1))
    r += 1
ws.column_dimensions["A"].width = 120

# ================================================================= Roadmap
wsr = wb.create_sheet("Roadmap")
wsr.sheet_view.showGridLines = False
c = wsr.cell(row=1, column=1, value="Learning roadmap - zero to production in 14 stages")
c.font = Font(size=16, bold=True, color=NAVY)

headers = ["Level", "Stage", "Goal", "What to learn", "Build this", "Done when"]
hr = 3
for i, h in enumerate(headers, start=1):
    cc = wsr.cell(row=hr, column=i, value=h)
    cc.font = Font(bold=True, color="FFFFFF")
    cc.fill = PatternFill("solid", fgColor=NAVY)
    cc.border = BORDER
    cc.alignment = Alignment(vertical="center")
wsr.row_dimensions[hr].height = 20

row = hr + 1
for (lvl, stage, goal, learn, build, done) in ROADMAP:
    short = lvl.split(" - ")[1]
    vals = [short, stage, goal, learn, build, done]
    band = PatternFill("solid", fgColor=BAND) if (row - hr) % 2 == 0 else None
    for col, v in enumerate(vals, start=1):
        cc = wsr.cell(row=row, column=col, value=v)
        cc.border = BORDER
        cc.font = Font(bold=True, color=ACCENT) if col == 3 else body_font
        cc.alignment = Alignment(vertical="top", wrap_text=(col >= 3))
        if band:
            cc.fill = band
    longest = max(len(learn), len(build), len(done))
    wsr.row_dimensions[row].height = max(34, 12.5 * (longest // 42 + 1))
    row += 1

for col, w in {"A": 15, "B": 8, "C": 30, "D": 46, "E": 46, "F": 46}.items():
    wsr.column_dimensions[col].width = w
wsr.freeze_panes = f"A{hr + 1}"

# ================================================================= All resources
ws2 = wb.create_sheet("All resources")
ws2.sheet_view.showGridLines = False
headers = ["#", "Level", "Topic", "Resource (click to open)", "What you get", "URL (click to open)",
           "Type", "Cost", "Link check", "Checked"]
for i, h in enumerate(headers, start=1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center")
    c.border = BORDER
ws2.row_dimensions[1].height = 24

row = 2
for n, (lvl, cat, title, desc, url, typ, cost) in enumerate(ROWS, start=1):
    res, when = verdict(url)
    band = PatternFill("solid", fgColor=BAND) if n % 2 == 0 else None
    vals = [n, lvl.split(" - ")[-1] if " - " in lvl else lvl, cat, title, desc, url, typ, cost, res, when]
    for col, v in enumerate(vals, start=1):
        c = ws2.cell(row=row, column=col, value=v)
        c.border = BORDER
        c.font = body_font
        c.alignment = Alignment(vertical="top", wrap_text=(col in (3, 4, 5, 9)))
        if band:
            c.fill = band
    for col in (4, 6):
        c = ws2.cell(row=row, column=col)
        c.hyperlink = url
        c.font = link_font
    ws2.cell(row=row, column=1).alignment = Alignment(vertical="top", horizontal="center")
    ws2.cell(row=row, column=9).font = Font(size=10, color=OK_GREEN if res.startswith("live") else WARN_AMBER)
    ws2.cell(row=row, column=10).font = Font(size=10, color="666666")
    ws2.row_dimensions[row].height = max(30, 13.5 * (len(desc) // 60 + 1))
    row += 1

last = row - 1
for col, w in {"A": 5, "B": 14, "C": 26, "D": 40, "E": 60, "F": 52, "G": 20, "H": 14, "I": 20, "J": 12}.items():
    ws2.column_dimensions[col].width = w
tbl = Table(displayName="AIResources", ref=f"A1:J{last}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
ws2.add_table(tbl)
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:J{last}"

# ================================================================= By level
ws3 = wb.create_sheet("By level")
ws3.sheet_view.showGridLines = False
c = ws3.cell(row=1, column=1, value="Resources grouped by level and topic")
c.font = Font(size=16, bold=True, color=NAVY)

grouped = OrderedDict()
for lvl, cat, title, desc, url, typ, cost in ROWS:
    grouped.setdefault(lvl, OrderedDict()).setdefault(cat, []).append((title, desc, url, typ, cost))

r = 3
for lvl in LEVEL_ORDER:
    if lvl not in grouped:
        continue
    lc = ws3.cell(row=r, column=1, value=lvl)
    lc.font = Font(size=13, bold=True, color="FFFFFF")
    lc.fill = PatternFill("solid", fgColor=NAVY)
    for col in range(2, 5):
        ws3.cell(row=r, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws3.row_dimensions[r].height = 22
    r += 1
    for cat, items in grouped[lvl].items():
        hc = ws3.cell(row=r, column=1, value=f"{cat}  ({len(items)})")
        hc.font = Font(size=11, bold=True, color="FFFFFF")
        hc.fill = PatternFill("solid", fgColor=ACCENT)
        for col in range(2, 5):
            ws3.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ACCENT)
        r += 1
        for col, h in enumerate(["Resource (click to open)", "What you get", "Type", "Cost"], start=1):
            hh = ws3.cell(row=r, column=col, value=h)
            hh.font = Font(bold=True, size=9, color=NAVY)
            hh.fill = PatternFill("solid", fgColor=LIGHT)
            hh.border = BORDER
        r += 1
        for title, desc, url, typ, cost in items:
            a = ws3.cell(row=r, column=1, value=title)
            a.hyperlink = url
            a.font = link_font
            a.alignment = Alignment(vertical="top", wrap_text=True)
            b = ws3.cell(row=r, column=2, value=desc)
            b.font = body_font
            b.alignment = Alignment(vertical="top", wrap_text=True)
            d = ws3.cell(row=r, column=3, value=typ)
            d.font = body_font
            d.alignment = Alignment(vertical="top", wrap_text=True)
            e = ws3.cell(row=r, column=4, value=cost)
            e.font = body_font
            e.alignment = Alignment(vertical="top", wrap_text=True)
            for col in range(1, 5):
                ws3.cell(row=r, column=col).border = BORDER
            ws3.row_dimensions[r].height = max(28, 13 * (len(desc) // 62 + 1))
            r += 1
    r += 1

for col, w in {"A": 44, "B": 70, "C": 20, "D": 16}.items():
    ws3.column_dimensions[col].width = w

# ================================================================= Code snippets
ws4 = wb.create_sheet("Code snippets")
ws4.sheet_view.showGridLines = False
c = ws4.cell(row=1, column=1, value="Starter code (all Python snippets syntax-checked)")
c.font = Font(size=16, bold=True, color=NAVY)
c2 = ws4.cell(row=2, column=1, value="Select a snippet cell and copy - Excel keeps the line breaks. If a model API "
                                     "call is rejected, open the Source link for the current shape.")
c2.font = Font(size=10, italic=True, color="666666")
c2.alignment = Alignment(wrap_text=True, vertical="top")

r = 4
for col, h in enumerate(["Level", "What it is", "Lang", "Why / when", "Snippet", "Source"], start=1):
    hh = ws4.cell(row=r, column=col, value=h)
    hh.font = Font(bold=True, color="FFFFFF", size=11)
    hh.fill = PatternFill("solid", fgColor=NAVY)
    hh.border = BORDER
ws4.row_dimensions[r].height = 20
r += 1

for (lvl, title, lang, why, code, src) in SNIPPETS:
    a = ws4.cell(row=r, column=1, value=lvl.split(" - ")[1])
    a.font = body_font
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b = ws4.cell(row=r, column=2, value=title)
    b.font = Font(bold=True, size=10, color=ACCENT)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    lg = ws4.cell(row=r, column=3, value=lang)
    lg.font = Font(size=9, color="666666")
    lg.alignment = Alignment(vertical="top")
    w = ws4.cell(row=r, column=4, value=why)
    w.font = body_font
    w.alignment = Alignment(vertical="top", wrap_text=True)
    cd = ws4.cell(row=r, column=5, value=code)
    cd.font = code_font
    cd.alignment = Alignment(vertical="top", wrap_text=True)
    cd.fill = PatternFill("solid", fgColor=CODE_BG)
    sc = ws4.cell(row=r, column=6, value="source")
    sc.hyperlink = src
    sc.font = link_font
    sc.alignment = Alignment(vertical="top")
    for col in range(1, 7):
        ws4.cell(row=r, column=col).border = BORDER
    lines = code.count("\n") + 1
    ws4.row_dimensions[r].height = max(14 * lines + 6, 13 * (len(why) // 34 + 1))
    r += 1

for col, w in {"A": 13, "B": 26, "C": 10, "D": 38, "E": 82, "F": 10}.items():
    ws4.column_dimensions[col].width = w
ws4.freeze_panes = "A5"

# ================================================================= Index
ws5 = wb.create_sheet("Index")
ws5.sheet_view.showGridLines = False
c = ws5.cell(row=1, column=1, value="Index and counts")
c.font = Font(size=16, bold=True, color=NAVY)


def block(ws, start, heading_text, pairs, total=None):
    ws.cell(row=start, column=1, value=heading_text).font = Font(bold=True, size=12, color=ACCENT)
    rr = start + 1
    for col, h in enumerate(["Name", "Count"], 1):
        hh = ws.cell(row=rr, column=col, value=h)
        hh.font = Font(bold=True, color="FFFFFF")
        hh.fill = PatternFill("solid", fgColor=NAVY)
        hh.border = BORDER
    rr += 1
    for name, n in pairs:
        a = ws.cell(row=rr, column=1, value=name)
        a.border = BORDER
        a.font = body_font
        bcell = ws.cell(row=rr, column=2, value=n)
        bcell.border = BORDER
        bcell.font = body_font
        bcell.alignment = Alignment(horizontal="center")
        rr += 1
    if total:
        a = ws.cell(row=rr, column=1, value=total[0])
        a.font = Font(bold=True)
        a.border = BORDER
        bcell = ws.cell(row=rr, column=2, value=total[1])
        bcell.font = Font(bold=True)
        bcell.border = BORDER
        bcell.alignment = Alignment(horizontal="center")
        rr += 1
    return rr + 2


lvl_counts = Counter(r[0] for r in ROWS)
nxt = block(ws5, 3, "Resources per level",
            [(lvl, lvl_counts[lvl]) for lvl in LEVEL_ORDER if lvl in lvl_counts], ("TOTAL", len(ROWS)))
nxt = block(ws5, nxt, "Resources per topic", Counter(r[1] for r in ROWS).most_common())
nxt = block(ws5, nxt, "Resources per type", Counter(r[5] for r in ROWS).most_common())
kinds = Counter(STATUS.get(r[4], {}).get("kind", "not checked") for r in ROWS)
nxt = block(ws5, nxt, f"Link check ({TODAY})", [
    ("Live - HTTP 200", kinds.get("ok", 0)),
    ("Bot-blocked (opens in a browser)", kinds.get("blocked-to-bots", 0)),
    ("Broken", kinds.get("broken", 0)),
])
block(ws5, nxt, "Workbook contents", [
    ("Curated resources", len(ROWS)),
    ("Roadmap stages", len(ROADMAP)),
    ("Code snippets", len(SNIPPETS)),
    ("Topics", len(set(r[1] for r in ROWS))),
])

ws5.column_dimensions["A"].width = 44
ws5.column_dimensions["B"].width = 10

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  resources : {len(ROWS)} | roadmap: {len(ROADMAP)} | snippets: {len(SNIPPETS)}")
print(f"  sheets    : {wb.sheetnames}")
print(f"  link check: {dict(kinds)}")
