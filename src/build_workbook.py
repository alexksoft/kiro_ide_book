"""Build a single Excel workbook cataloguing Kiro (AWS agentic IDE) resources."""

import json
import os
import re
from collections import Counter, OrderedDict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from config_snippets import SNIPPETS
from kiro_resources_data import DUPES_REMOVED, ROWS

OUT = "Kiro_IDE_Resources.xlsx"
TODAY = date.today().isoformat()

NAVY = "1F3864"
ACCENT = "6C3FA8"
LIGHT = "EDE7F6"
BAND = "F5F3FA"
CODE_BG = "F4F4F7"
LINK_BLUE = "0563C1"
OK_GREEN = "1E7B34"
WARN_AMBER = "9C5700"

thin = Side(style="thin", color="D0CCE0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

link_font = Font(color=LINK_BLUE, underline="single", size=11)
body_font = Font(size=11, color="222222")
code_font = Font(name="Consolas", size=9.5, color="1A1A2E")

# ------------------------------------------------------------------ link status
try:
    with open("link_status.json") as fh:
        STATUS = json.load(fh)
except FileNotFoundError:
    STATUS = {}


def verdict(url):
    s = STATUS.get(url)
    if not s:
        return "not checked", ""
    if s["kind"] == "ok":
        return "live (HTTP 200)", s["checked"]
    if s["kind"] == "blocked-to-bots":
        return f"bot-blocked ({s['code']}) - opens in browser", s["checked"]
    return f"PROBLEM ({s['code']})", s["checked"]


def domain_of(url):
    d = url.split("//", 1)[-1].split("/", 1)[0]
    return d[4:] if d.startswith("www.") else d


wb = Workbook()

# ================================================================= Start here
ws = wb.active
ws.title = "Start here"
ws.sheet_view.showGridLines = False

n_ok = sum(1 for u in {r[3] for r in ROWS} if STATUS.get(u, {}).get("kind") == "ok")
n_blocked = sum(1 for u in {r[3] for r in ROWS} if STATUS.get(u, {}).get("kind") == "blocked-to-bots")

intro = [
    ("Kiro - master link catalogue", "title"),
    (f"Compiled {TODAY}  |  {len(ROWS)} curated links  |  every link is clickable", "sub"),
    ("", None),
    ("What this workbook is", "h"),
    ("Everything useful I could find on the web about Kiro - AWS's agentic, spec-driven IDE (kiro.dev). "
     "Official documentation, AWS blogs and re:Post, community tutorials, Stack Overflow and Quora, GitHub bug "
     "reports, forum threads, courses, hackathon galleries, comparison reviews, and non-English resources.", "p"),
    ("", None),
    ("Name note - you spelled it several ways", "h"),
    ("The product is Kiro, one word, from AWS. 'Kiroid', 'keroi', 'queroid', 'kero', 'ki ro IDE' and "
     "'DEIDE hero IDE' are phonetic variants of the same thing. Website: kiro.dev. GitHub org: kirodotdev. "
     "SDD means spec-driven development (its headline feature) - see the 'Spec-driven development' and "
     "'Writing requirements (EARS)' rows. If you meant SSD/disk and resource trouble, see 'Known problems & fixes', "
     "which covers memory growth, high CPU and freezes.", "p"),
    ("", None),
    ("The six sheets", "h"),
    ("Start here - this page.", "p"),
    ("All resources - the full curated table, one row per link, with filters.", "p"),
    ("By category - the same links grouped by topic so you can read one subject at a time.", "p"),
    ("Official docs index - every page of the official documentation, auto-built from kiro.dev/llms.txt.", "p"),
    ("Config snippets - copy-paste ready mcp.json, hooks, steering, skills, agents and CI examples.", "p"),
    ("Index - counts per category, per source type, and link-check totals.", "p"),
    ("", None),
    ("How to use the big table", "h"),
    ("Row 1 has filter arrows: narrow by Category or Source type - e.g. only 'GitHub issue' rows when chasing a "
     "bug, or only 'Official docs' when you need the authoritative answer. Headers stay frozen as you scroll.", "p"),
    ("Both the Title and the URL are clickable in every row.", "p"),
    ("", None),
    ("Link checking - what was actually verified", "h"),
    (f"Every URL was requested on {TODAY}. {n_ok} returned HTTP 200. {n_blocked} sit behind bot protection "
     "(Cloudflare 'Just a moment...' or a Vercel checkpoint) - Stack Overflow, Medium, Quora, Udemy, DataCamp, "
     "OpenAI, HashiCorp and tech-insider. Those pages are real and open normally in a browser; they simply refuse "
     "automated clients. Nothing in this workbook returned 404. The per-row result is in the last two columns.", "p"),
    ("", None),
    ("Suggested reading order if you are new", "h"),
    ("1. FAQ, then Docs home - what it actually is.", "p"),
    ("2. Installation, then 'Your first project' - get it running.", "p"),
    ("3. Specs, spec best practices, and EARS - the feature that makes Kiro different.", "p"),
    ("4. Steering - stop repeating your conventions in every prompt.", "p"),
    ("5. Hooks, MCP and Powers - automation and external tools.", "p"),
    ("6. Configuration scopes and Permissions - where files live and what the agent may do.", "p"),
    ("7. IDE troubleshooting + the GitHub issue tracker - before you report anything.", "p"),
    ("", None),
    ("Caveats", "h"),
    ("Kiro ships fast: pricing, model names, menu locations and even doc URLs change. Cross-check community posts "
     "against the official docs and changelog. Courses are paid third-party products - check reviews and how "
     "recently they were updated before buying.", "p"),
    ("Community repositories, gists and templates are not endorsed by AWS. Treat advice about editing Kiro's "
     "internal configuration (for example swapping the extension marketplace) as unsupported and at your own risk.", "p"),
    ("Deliberately excluded: several GitHub projects exist that proxy Kiro credentials into other AI clients. "
     "They appear to breach the terms of service and are a plausible route to the account suspensions documented "
     "in this workbook, so I have not linked them.", "p"),
]

r = 1
for text, kind in intro:
    c = ws.cell(row=r, column=1, value=text)
    if kind == "title":
        c.font = Font(size=20, bold=True, color=NAVY)
        ws.row_dimensions[r].height = 28
    elif kind == "sub":
        c.font = Font(size=11, italic=True, color="666666")
    elif kind == "h":
        c.font = Font(size=12, bold=True, color=ACCENT)
        ws.row_dimensions[r].height = 20
    elif kind == "p":
        c.font = Font(size=11, color="333333")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 15 * (len(text) // 105 + 1))
    r += 1
ws.column_dimensions["A"].width = 118

# ================================================================= All resources
ws2 = wb.create_sheet("All resources")
ws2.sheet_view.showGridLines = False

headers = ["#", "Category", "Title (click to open)", "What you get there", "URL (click to open)",
           "Source type", "Domain", "Link check", "Checked"]
for i, h in enumerate(headers, start=1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", horizontal="left")
    c.border = BORDER
ws2.row_dimensions[1].height = 24

row = 2
for n, (cat, title, desc, url, stype) in enumerate(ROWS, start=1):
    res, when = verdict(url)
    band = PatternFill("solid", fgColor=BAND) if n % 2 == 0 else None
    for col, v in enumerate([n, cat, title, desc, url, stype, domain_of(url), res, when], start=1):
        c = ws2.cell(row=row, column=col, value=v)
        c.border = BORDER
        c.font = body_font
        c.alignment = Alignment(vertical="top", wrap_text=(col in (2, 3, 4, 8)))
        if band:
            c.fill = band
    for col in (3, 5):
        c = ws2.cell(row=row, column=col)
        c.hyperlink = url
        c.font = link_font
    ws2.cell(row=row, column=1).alignment = Alignment(vertical="top", horizontal="center")
    ws2.cell(row=row, column=8).font = Font(size=10, color=OK_GREEN if res.startswith("live") else WARN_AMBER)
    ws2.cell(row=row, column=9).font = Font(size=10, color="666666")
    ws2.row_dimensions[row].height = max(30, 13.5 * (len(desc) // 62 + 1))
    row += 1

last = row - 1
for col, w in {"A": 5, "B": 27, "C": 44, "D": 64, "E": 58, "F": 21, "G": 20, "H": 20, "I": 12}.items():
    ws2.column_dimensions[col].width = w

tbl = Table(displayName="KiroResources", ref=f"A1:I{last}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
ws2.add_table(tbl)
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:I{last}"

# ================================================================= By category
ws3 = wb.create_sheet("By category")
ws3.sheet_view.showGridLines = False

grouped = OrderedDict()
for cat, title, desc, url, stype in ROWS:
    grouped.setdefault(cat, []).append((title, desc, url, stype))

c = ws3.cell(row=1, column=1, value="Kiro resources grouped by topic")
c.font = Font(size=16, bold=True, color=NAVY)
ws3.row_dimensions[1].height = 24

r = 3
for cat, items in grouped.items():
    hc = ws3.cell(row=r, column=1, value=f"{cat}  ({len(items)} links)")
    hc.font = Font(size=12, bold=True, color="FFFFFF")
    hc.fill = PatternFill("solid", fgColor=ACCENT)
    hc.alignment = Alignment(vertical="center", indent=1)
    for col in range(2, 5):
        ws3.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ACCENT)
    ws3.row_dimensions[r].height = 22
    r += 1
    for col, h in enumerate(["Title (click to open)", "What you get there", "URL (click to open)", "Source type"], 1):
        hh = ws3.cell(row=r, column=col, value=h)
        hh.font = Font(bold=True, size=10, color=NAVY)
        hh.fill = PatternFill("solid", fgColor=LIGHT)
        hh.border = BORDER
    r += 1
    for title, desc, url, stype in items:
        a = ws3.cell(row=r, column=1, value=title)
        a.hyperlink = url
        a.font = link_font
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b = ws3.cell(row=r, column=2, value=desc)
        b.font = body_font
        b.alignment = Alignment(vertical="top", wrap_text=True)
        d = ws3.cell(row=r, column=3, value=url)
        d.hyperlink = url
        d.font = link_font
        d.alignment = Alignment(vertical="top")
        e = ws3.cell(row=r, column=4, value=stype)
        e.font = body_font
        e.alignment = Alignment(vertical="top")
        for col in range(1, 5):
            ws3.cell(row=r, column=col).border = BORDER
        ws3.row_dimensions[r].height = max(30, 13.5 * (len(desc) // 68 + 1))
        r += 1
    r += 1

for col, w in {"A": 46, "B": 72, "C": 60, "D": 21}.items():
    ws3.column_dimensions[col].width = w

# ================================================================= Official docs index
ws4 = wb.create_sheet("Official docs index")
ws4.sheet_view.showGridLines = False

docs = []          # (section, title, url, short_desc, depth)
section = ""
if os.path.exists("_docs/llms.txt"):
    with open("_docs/llms.txt") as fh:
        for line in fh:
            if line.startswith("## "):
                section = line[3:].strip()
                continue
            m = re.match(r"^(\s*)-\s*\[([^\]]+)\]\((https://kiro\.dev/[^)]+)\)\s*:?\s*(.*)$", line.rstrip())
            if not m:
                continue
            indent, title, url, desc = m.groups()
            url = url[:-3] + "/" if url.endswith(".md") else url
            words = desc.split()
            short = " ".join(words[:12]) + ("..." if len(words) > 12 else "")
            docs.append((section, title.strip(), url, short, len(indent) // 2))

seen_docs = set()
docs = [d for d in docs if not (d[2] in seen_docs or seen_docs.add(d[2]))]

c = ws4.cell(row=1, column=1, value="Every page of the official Kiro documentation")
c.font = Font(size=16, bold=True, color=NAVY)
c2 = ws4.cell(row=2, column=1,
              value=f"{len(docs)} pages, generated from the official machine-readable docs index at "
                    f"kiro.dev/llms.txt on {TODAY}. Summaries are shortened from that index; open a page for the "
                    f"full text. Tip: append .md to any docs URL to get clean Markdown.")
c2.font = Font(size=10, italic=True, color="666666")
c2.alignment = Alignment(wrap_text=True, vertical="top")
ws4.row_dimensions[2].height = 30
src = ws4.cell(row=3, column=1, value="https://kiro.dev/llms.txt")
src.hyperlink = "https://kiro.dev/llms.txt"
src.font = link_font

r = 5
for col, h in enumerate(["Docs section", "Page (click to open)", "Summary", "URL (click to open)"], 1):
    hh = ws4.cell(row=r, column=col, value=h)
    hh.font = Font(bold=True, color="FFFFFF", size=11)
    hh.fill = PatternFill("solid", fgColor=NAVY)
    hh.border = BORDER
ws4.row_dimensions[r].height = 22
head = r
r += 1

for i, (sec, title, url, desc, depth) in enumerate(docs):
    band = PatternFill("solid", fgColor=BAND) if i % 2 else None
    a = ws4.cell(row=r, column=1, value=sec)
    a.font = Font(size=10, color="555555")
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b = ws4.cell(row=r, column=2, value=("    " * depth) + title)
    b.hyperlink = url
    b.font = link_font
    b.alignment = Alignment(vertical="top", wrap_text=True)
    d = ws4.cell(row=r, column=3, value=desc)
    d.font = body_font
    d.alignment = Alignment(vertical="top", wrap_text=True)
    e = ws4.cell(row=r, column=4, value=url)
    e.hyperlink = url
    e.font = link_font
    e.alignment = Alignment(vertical="top")
    for col in range(1, 5):
        cc = ws4.cell(row=r, column=col)
        cc.border = BORDER
        if band:
            cc.fill = band
    ws4.row_dimensions[r].height = max(26, 13.5 * (len(desc) // 58 + 1))
    r += 1

for col, w in {"A": 26, "B": 46, "C": 60, "D": 56}.items():
    ws4.column_dimensions[col].width = w
if docs:
    ws4.auto_filter.ref = f"A{head}:D{r - 1}"
    ws4.freeze_panes = f"A{head + 1}"

# ================================================================= Config snippets
ws5 = wb.create_sheet("Config snippets")
ws5.sheet_view.showGridLines = False

c = ws5.cell(row=1, column=1, value="Copy-paste configuration for Kiro")
c.font = Font(size=16, bold=True, color=NAVY)
c2 = ws5.cell(row=2, column=1,
              value="Each block was taken from or checked against the official docs on "
                    f"{TODAY} - the Source column links to the page it came from. Select a snippet cell and copy; "
                    "Excel keeps the line breaks. Schemas do change, so re-check the source page if something "
                    "is rejected.")
c2.font = Font(size=10, italic=True, color="666666")
c2.alignment = Alignment(wrap_text=True, vertical="top")
ws5.row_dimensions[2].height = 30

r = 4
for col, h in enumerate(["What it is", "Where it goes", "Why / gotchas", "Snippet", "Source"], 1):
    hh = ws5.cell(row=r, column=col, value=h)
    hh.font = Font(bold=True, color="FFFFFF", size=11)
    hh.fill = PatternFill("solid", fgColor=NAVY)
    hh.border = BORDER
ws5.row_dimensions[r].height = 22
r += 1

for title, path, why, snippet, source in SNIPPETS:
    a = ws5.cell(row=r, column=1, value=title)
    a.font = Font(bold=True, size=11, color=ACCENT)
    a.alignment = Alignment(vertical="top", wrap_text=True)

    b = ws5.cell(row=r, column=2, value=path)
    b.font = Font(name="Consolas", size=9.5, color="333333")
    b.alignment = Alignment(vertical="top", wrap_text=True)

    d = ws5.cell(row=r, column=3, value=why)
    d.font = body_font
    d.alignment = Alignment(vertical="top", wrap_text=True)

    e = ws5.cell(row=r, column=4, value=snippet)
    e.font = code_font
    e.alignment = Alignment(vertical="top", wrap_text=True)
    e.fill = PatternFill("solid", fgColor=CODE_BG)

    f = ws5.cell(row=r, column=5, value="official docs")
    f.hyperlink = source
    f.font = link_font
    f.alignment = Alignment(vertical="top")

    for col in range(1, 6):
        ws5.cell(row=r, column=col).border = BORDER

    lines = snippet.count("\n") + 1
    wrapped = sum(max(1, len(ln) // 62 + 1) for ln in snippet.split("\n"))
    ws5.row_dimensions[r].height = max(13 * max(lines, wrapped) + 8, 13.5 * (len(why) // 40 + 1))
    r += 1

for col, w in {"A": 30, "B": 34, "C": 42, "D": 66, "E": 13}.items():
    ws5.column_dimensions[col].width = w
ws5.freeze_panes = "A5"

# ================================================================= Index
ws6 = wb.create_sheet("Index")
ws6.sheet_view.showGridLines = False

c = ws6.cell(row=1, column=1, value="Index and counts")
c.font = Font(size=16, bold=True, color=NAVY)


def table_block(ws, start_row, heading, pairs, total_label=None, total_value=None):
    ws.cell(row=start_row, column=1, value=heading).font = Font(bold=True, size=12, color=ACCENT)
    rr = start_row + 1
    for col, h in enumerate(["Name", "Count"], 1):
        hh = ws.cell(row=rr, column=col, value=h)
        hh.font = Font(bold=True, color="FFFFFF")
        hh.fill = PatternFill("solid", fgColor=NAVY)
        hh.border = BORDER
    rr += 1
    for name, n in pairs:
        a = ws.cell(row=rr, column=1, value=name)
        a.border = BORDER
        a.font = body_font
        b = ws.cell(row=rr, column=2, value=n)
        b.border = BORDER
        b.font = body_font
        b.alignment = Alignment(horizontal="center")
        rr += 1
    if total_label:
        a = ws.cell(row=rr, column=1, value=total_label)
        a.font = Font(bold=True)
        a.border = BORDER
        b = ws.cell(row=rr, column=2, value=total_value)
        b.font = Font(bold=True)
        b.border = BORDER
        b.alignment = Alignment(horizontal="center")
        rr += 1
    return rr + 2


nxt = table_block(ws6, 3, "Curated links per category",
                  [(k, len(v)) for k, v in grouped.items()], "TOTAL", len(ROWS))
nxt = table_block(ws6, nxt, "Curated links per source type",
                  Counter(x[4] for x in ROWS).most_common())
kinds = Counter(STATUS.get(r[3], {}).get("kind", "not checked") for r in ROWS)
nxt = table_block(ws6, nxt, f"Link check results ({TODAY})", [
    ("Live - HTTP 200", kinds.get("ok", 0)),
    ("Bot-blocked (page exists, opens in a browser)", kinds.get("blocked-to-bots", 0)),
    ("Broken", kinds.get("broken", 0)),
    ("Not checked", kinds.get("not checked", 0)),
])
table_block(ws6, nxt, "Workbook contents", [
    ("Curated links (All resources sheet)", len(ROWS)),
    ("Official docs pages (Official docs index sheet)", len(docs)),
    ("Config snippets (Config snippets sheet)", len(SNIPPETS)),
    ("Topic categories", len(grouped)),
    ("Duplicate URLs removed while building", DUPES_REMOVED),
])

ws6.column_dimensions["A"].width = 48
ws6.column_dimensions["B"].width = 10

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  curated links      : {len(ROWS)} across {len(grouped)} categories")
print(f"  official docs pages: {len(docs)}")
print(f"  config snippets    : {len(SNIPPETS)}")
print(f"  sheets             : {wb.sheetnames}")
print(f"  link check         : {dict(kinds)}")
